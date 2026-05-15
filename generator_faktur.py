"""
Generator raportów faktur według księgowych
Загрузи CSV (lista faktur) и XLSX (lista klientów) — получи готовый Excel по бухгалтерам
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import re
import os
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ─────────────────────────────────────────
#  FILE READER — auto-detect format
# ─────────────────────────────────────────

def read_file(path, **kwargs):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return pd.read_excel(path, **kwargs)
    # CSV — пробуем cp1250, потом utf-8
    for enc in ('cp1250', 'utf-8', 'utf-8-sig'):
        try:
            return pd.read_csv(path, sep=';', encoding=enc, **kwargs)
        except (UnicodeDecodeError, ValueError):
            continue
    raise ValueError(f"Не удалось прочитать файл: {path}")


# ─────────────────────────────────────────
#  MATCHING
# ─────────────────────────────────────────

def extract_nip(k):
    m = re.search(r'NIP/PESEL:\s*(\d+)', str(k))
    return m.group(1) if m else ''

def extract_name(k):
    return re.sub(r'\s*NIP/PESEL:\s*\d+', '', str(k)).strip()

def normalize(name):
    name = str(name).upper().strip()
    name = re.sub(r'[\n\r]', ' ', name)
    for pat in [r'^\(JDG\s*IT\)\s*', r'^\(JDG\)\s*', r'^JDG\s+']:
        name = re.sub(pat, '', name)
    name = re.sub(r'[.,\-]', '', name)
    name = re.sub(r'SP[ÓO][Ł]KA\s+Z\s+OGRANICZON[AĄ]\s+ODPOWIEDZIALNO[SŚ]CI[AĄ]', 'SPZOO', name)
    name = re.sub(r'SP\s+Z\s+O\s+O\b', 'SPZOO', name)
    name = re.sub(r'SP\s+Z\s+OO\b', 'SPZOO', name)
    name = re.sub(r'\bPSA\b', 'SPZOO', name)
    return re.sub(r'\s+', ' ', name).strip()

def find_accountant(csv_name, xl_norm_list):
    csv_core = normalize(csv_name).replace('SPZOO', '').strip()
    best_score, best_rec = 0, None
    for xl_norm, rec in xl_norm_list:
        score = fuzz.WRatio(csv_core, xl_norm.replace('SPZOO', '').strip())
        if score > best_score:
            best_score, best_rec = score, rec
    if best_score >= 88:
        return best_rec['Buchhalter'], best_rec['Pomochnik']
    return '', ''


# ─────────────────────────────────────────
#  EXCEL GENERATION
# ─────────────────────────────────────────

def parse_amount(val):
    try:
        return float(re.sub(r'[^\d,]', '', str(val)).replace(',', '.'))
    except:
        return 0.0

ACCOUNTANT_COLORS = {
    'Aneta':                 {'header': '1F497D', 'subheader': '2E75B6'},
    'Dominika':              {'header': '375623', 'subheader': '70AD47'},
    'Angelika Olejniczak':   {'header': 'BE4B05', 'subheader': 'ED7D31'},
    'Jadwiga':               {'header': '7F6000', 'subheader': 'FFC000'},
    'Iwona':                 {'header': '3D1F6B', 'subheader': '7030A0'},
    'Agnieszka Orczykowska': {'header': '820000', 'subheader': 'C00000'},
    'Anieszka Młynarska':    {'header': '005E7C', 'subheader': '00B0F0'},
}
DEFAULT_COLOR = {'header': '404040', 'subheader': '808080'}

thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

INV_HEADERS = [
    'Lp', 'Klient (pełna nazwa)', 'NIP', 'Numer faktury',
    'Data wystawienia', 'Termin płatności',
    'Kwota netto', 'Kwota brutto', 'Waluta',
    'Zapłacona', 'Zatwierdzona', 'KSeF', 'Pomochnik/Asystent'
]
COL_WIDTHS = [5, 42, 13, 16, 17, 17, 15, 15, 8, 12, 14, 14, 22]


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def write_header_row(ws, row_num, values, bg_color, font_color='FFFFFF', font_size=10, height=20):
    fill = PatternFill("solid", fgColor=bg_color)
    font = Font(bold=True, color=font_color, size=font_size)
    ws.row_dimensions[row_num].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

def write_invoice_rows(ws, grp, start_row):
    alt_fill = PatternFill("solid", fgColor="EEF2F7")
    for i, (_, row_data) in enumerate(grp.iterrows()):
        r = start_row + i
        netto_str = str(row_data.get('Kwota netto', ''))
        curr_match = re.search(r'[A-Z]{3}', netto_str)
        currency = curr_match.group() if curr_match else 'PLN'
        netto_val  = parse_amount(netto_str)
        brutto_val = parse_amount(str(row_data.get('Kwota brutto', '')))
        vals = [
            row_data.get('Lp', ''),
            row_data.get('Klient', ''),
            str(row_data.get('NIP', '')),
            row_data.get('Numer', ''),
            row_data.get('Data wystawienia', ''),
            row_data.get('Termin płatności', ''),
            netto_val,
            brutto_val,
            currency,
            row_data.get('Zapłacona', ''),
            row_data.get('Zatwierdzona', ''),
            row_data.get('KSeF', ''),
            row_data.get('Pomochnik', ''),
        ]
        ws.row_dimensions[r].height = 16
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if i % 2 == 1:
                c.fill = alt_fill
            c.alignment = Alignment(vertical='center', wrap_text=(col == 2))
            c.border = BORDER
            if col in (7, 8):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            elif col in (1, 9, 10, 11, 12):
                c.alignment = Alignment(horizontal='center', vertical='center')


def generate_excel(df_inv, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    matched_df  = df_inv[df_inv['Buchhalter'] != ''].copy()
    not_found_df = df_inv[df_inv['Buchhalter'] == ''].copy()
    accountants  = sorted(matched_df['Buchhalter'].unique())
    ncols        = len(INV_HEADERS)
    last_col_letter = get_column_letter(ncols)

    # словарь: бухгалтер → строка RAZEM на его листе (для формул PODSUMOWANIE)
    razem_refs = {}

    # ── Листы по бухгалтерам (сначала, чтобы PODSUMOWANIE мог ссылаться) ──
    for buch in accountants:
        grp  = matched_df[matched_df['Buchhalter'] == buch].copy()
        clr  = ACCOUNTANT_COLORS.get(buch, DEFAULT_COLOR)
        safe = re.sub(r'[^\w]', '_', buch)[:28]   # имя таблицы без спецсимволов
        ws   = wb.create_sheet(buch[:31])
        ws.sheet_view.showGridLines = False

        # ── Заголовок-баннер (строка 1) ──
        ws.merge_cells(f'A1:{last_col_letter}1')
        c = ws['A1']
        c.value = f"Księgowy: {buch}   |   Liczba faktur: {len(grp)}   |   ⬇ Filtruj kolumny aby zmienić sumy"
        c.font  = Font(bold=True, size=12, color='FFFFFF')
        c.fill  = PatternFill("solid", fgColor=clr['header'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        # ── Заголовки (строка 2) ──
        write_header_row(ws, 2, INV_HEADERS, clr['subheader'], height=22)
        set_col_widths(ws, COL_WIDTHS)

        # ── Данные (строки 3…) ──
        write_invoice_rows(ws, grp, 3)
        last_data_row = 2 + len(grp)   # последняя строка данных включая заголовок

        # ── Excel Table с автофильтром ──
        tbl_ref  = f"A2:{last_col_letter}{last_data_row}"
        tbl_name = f"Tbl_{safe}"
        tbl = Table(displayName=tbl_name, ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,  showColumnStripes=False)
        ws.add_table(tbl)

        # ── Строка RAZEM с SUBTOTAL (пересчитывается по фильтру) ──
        total_row = last_data_row + 1
        razem_refs[buch] = (ws.title, total_row)
        ws.row_dimensions[total_row].height = 20
        total_fill = PatternFill("solid", fgColor=clr['subheader'])

        data_start = 3
        data_end   = last_data_row

        for col in range(1, ncols + 1):
            c = ws.cell(row=total_row, column=col)
            c.fill  = total_fill
            c.font  = Font(bold=True, color='FFFFFF', size=11)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')

        # RAZEM label
        ws.cell(row=total_row, column=1, value='RAZEM (widoczne)')

        # SUBTOTAL(9,...) = SUM видимых строк; col G=7 netto, H=8 brutto
        col_g = get_column_letter(7)
        col_h = get_column_letter(8)
        col_a = get_column_letter(1)

        c7 = ws.cell(row=total_row, column=7)
        c7.value  = f"=SUBTOTAL(9,{col_g}{data_start}:{col_g}{data_end})"
        c7.number_format = '#,##0.00'
        c7.alignment = Alignment(horizontal='right', vertical='center')

        c8 = ws.cell(row=total_row, column=8)
        c8.value  = f"=SUBTOTAL(9,{col_h}{data_start}:{col_h}{data_end})"
        c8.number_format = '#,##0.00'
        c8.alignment = Alignment(horizontal='right', vertical='center')

        # SUBTOTAL(103,...) = COUNT видимых строк
        c_cnt = ws.cell(row=total_row, column=2)
        c_cnt.value = f"=SUBTOTAL(103,{col_a}{data_start}:{col_a}{data_end})"
        c_cnt.alignment = Alignment(horizontal='center', vertical='center')

    # ── PODSUMOWANIE (ссылается на RAZEM каждого листа) ──
    ws_sum = wb.create_sheet("PODSUMOWANIE", 0)
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:F1')
    c = ws_sum['A1']
    c.value = 'PODSUMOWANIE — aktualizuje się automatycznie po filtrach'
    c.font  = Font(bold=True, size=13, color='FFFFFF')
    c.fill  = PatternFill("solid", fgColor='1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 28

    write_header_row(ws_sum, 2,
        ['Księgowy', 'Widocznych faktur', 'Kwota netto', 'Kwota brutto', 'Arkusz', ''],
        '2E4057', height=22)

    sum_row = 3
    for buch in accountants:
        clr = ACCOUNTANT_COLORS.get(buch, DEFAULT_COLOR)
        fill = PatternFill("solid", fgColor=clr['subheader'])
        sheet_title, tr = razem_refs[buch]
        safe_title = f"'{sheet_title}'" if ' ' in sheet_title else sheet_title

        cells = [
            (1, buch,   None),
            (2, f"={safe_title}!B{tr}", '#,##0'),
            (3, f"={safe_title}!G{tr}", '#,##0.00'),
            (4, f"={safe_title}!H{tr}", '#,##0.00'),
            (5, f'=HYPERLINK("#\'{sheet_title}\'!A1","→ перейти")', None),
        ]
        for col, val, fmt in cells:
            c = ws_sum.cell(row=sum_row, column=col, value=val)
            c.font  = Font(bold=(col == 1), color='FFFFFF', size=10)
            c.fill  = fill
            c.border = BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fmt:
                c.number_format = fmt
        ws_sum.row_dimensions[sum_row].height = 18
        sum_row += 1

    # NIE ZNALEZIONO строка
    for col, val in enumerate(['NIE ZNALEZIONO', len(not_found_df), '-', '-', '-', ''], 1):
        c = ws_sum.cell(row=sum_row, column=col, value=val)
        c.font  = Font(bold=True, color='FFFFFF', size=10)
        c.fill  = PatternFill("solid", fgColor='C00000')
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[sum_row].height = 18

    set_col_widths(ws_sum, [28, 18, 18, 18, 14, 8])

    # ── NIE ZNALEZIONO лист ──
    ws_nf = wb.create_sheet("NIE ZNALEZIONO")
    ws_nf.sheet_view.showGridLines = False
    ws_nf.merge_cells(f'A1:{last_col_letter}1')
    c = ws_nf['A1']
    c.value = f"Faktury bez przypisanego księgowego — {len(not_found_df)} pozycji"
    c.font  = Font(bold=True, size=12, color='FFFFFF')
    c.fill  = PatternFill("solid", fgColor='7B0000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_nf.row_dimensions[1].height = 26
    write_header_row(ws_nf, 2, INV_HEADERS, 'C00000', height=22)
    set_col_widths(ws_nf, COL_WIDTHS)
    write_invoice_rows(ws_nf, not_found_df, 3)

    wb.save(output_path)
    return len(matched_df), len(not_found_df), accountants


# ─────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Generator raportów faktur")
        self.root.geometry("520x340")
        self.root.resizable(False, False)
        self.root.configure(bg='#F0F4F8')

        self.csv_path  = tk.StringVar(value='')
        self.xlsx_path = tk.StringVar(value='')

        self._build_ui()

    def _build_ui(self):
        bg       = '#E8EDF2'
        btn_bg   = '#1A5CA8'
        btn_fg   = '#FFFFFF'
        btn_act  = '#0D3E7A'
        run_bg   = '#0D3E7A'
        run_act  = '#092D5C'
        lbl_fg   = '#0A1F3C'

        # Title
        tk.Label(self.root, text="Generator raportów faktur",
                 font=('Helvetica', 14, 'bold'), bg=bg, fg='#1A5CA8').pack(pady=(18, 4))
        tk.Label(self.root, text="Załaduj dwa pliki — otrzymasz gotowy raport Excel",
                 font=('Helvetica', 10), bg=bg, fg='#333333').pack(pady=(0, 16))

        frame = tk.Frame(self.root, bg=bg)
        frame.pack(padx=30, fill='x')

        # File 1
        tk.Label(frame, text="1.  Lista faktur (.csv / .xlsx):",
                 font=('Helvetica', 11, 'bold'), bg=bg, fg=lbl_fg,
                 anchor='w').grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 4))
        tk.Entry(frame, textvariable=self.csv_path, width=34,
                 font=('Helvetica', 9), bg='#FFFFFF', fg='#111111',
                 relief='solid', bd=1).grid(row=1, column=0, sticky='ew', pady=(0, 8))
        tk.Button(frame, text="  Wybierz plik  ", command=self._pick_csv,
                  bg=btn_bg, fg=btn_fg, activebackground=btn_act, activeforeground=btn_fg,
                  relief='raised', bd=2, pady=4,
                  font=('Helvetica', 10, 'bold'), cursor='hand2').grid(row=1, column=1, padx=(8, 0), pady=(0, 8))

        # File 2
        tk.Label(frame, text="2.  Lista klientów z księgowymi (.csv / .xlsx):",
                 font=('Helvetica', 11, 'bold'), bg=bg, fg=lbl_fg,
                 anchor='w').grid(row=2, column=0, columnspan=2, sticky='w', pady=(4, 4))
        tk.Entry(frame, textvariable=self.xlsx_path, width=34,
                 font=('Helvetica', 9), bg='#FFFFFF', fg='#111111',
                 relief='solid', bd=1).grid(row=3, column=0, sticky='ew', pady=(0, 4))
        tk.Button(frame, text="  Wybierz plik  ", command=self._pick_xlsx,
                  bg=btn_bg, fg=btn_fg, activebackground=btn_act, activeforeground=btn_fg,
                  relief='raised', bd=2, pady=4,
                  font=('Helvetica', 10, 'bold'), cursor='hand2').grid(row=3, column=1, padx=(8, 0))

        frame.columnconfigure(0, weight=1)

        # Progress bar
        self.progress = ttk.Progressbar(self.root, mode='indeterminate', length=300)

        # Status
        self.status_var = tk.StringVar(value='')
        tk.Label(self.root, textvariable=self.status_var,
                 font=('Helvetica', 9), bg=bg, fg='#222222').pack(pady=(10, 0))

        # Generate button
        tk.Button(self.root, text="▶   Generuj raport",
                  command=self._run,
                  bg=run_bg, fg=btn_fg, activebackground=run_act, activeforeground=btn_fg,
                  font=('Helvetica', 12, 'bold'), relief='raised', bd=3,
                  padx=28, pady=12, cursor='hand2').pack(pady=(14, 0))

    def _pick_csv(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik z fakturami",
            filetypes=[("Wszystkie obsługiwane", "*.csv *.xlsx *.xls"),
                       ("CSV files", "*.csv"),
                       ("Excel files", "*.xlsx *.xls"),
                       ("All files", "*.*")])
        if path:
            self.csv_path.set(path)

    def _pick_xlsx(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik z klientami",
            filetypes=[("Wszystkie obsługiwane", "*.csv *.xlsx *.xls"),
                       ("CSV files", "*.csv"),
                       ("Excel files", "*.xlsx *.xls"),
                       ("All files", "*.*")])
        if path:
            self.xlsx_path.set(path)

    def _run(self):
        csv_path  = self.csv_path.get().strip()
        xlsx_path = self.xlsx_path.get().strip()

        if not csv_path or not os.path.exists(csv_path):
            messagebox.showerror("Błąd", "Wskaż plik z fakturami!")
            return
        if not xlsx_path or not os.path.exists(xlsx_path):
            messagebox.showerror("Błąd", "Wskaż plik z klientami!")
            return

        self.status_var.set("Przetwarzanie...")
        self.progress.pack(pady=4)
        self.progress.start(10)
        self.root.update()

        try:
            # Load invoices — auto-detect format
            df_inv = read_file(csv_path)
            df_inv = df_inv.drop(columns=[c for c in df_inv.columns if 'Unnamed' in c], errors='ignore')
            df_inv['NIP']    = df_inv['Kontrahent'].apply(extract_nip)
            df_inv['Klient'] = df_inv['Kontrahent'].apply(extract_name)

            # Load clients — auto-detect format
            df_xl = read_file(xlsx_path, header=None)
            df_xl.columns = ['Klient', 'Buchhalter', 'Pomochnik']
            df_xl = df_xl[df_xl['Klient'].notna()]
            df_xl = df_xl[~df_xl['Klient'].astype(str).str.contains('бухг', na=False)]
            df_xl['Klient']     = df_xl['Klient'].astype(str).str.strip().str.lstrip('\n').str.strip()
            df_xl['Buchhalter'] = df_xl['Buchhalter'].astype(str).str.strip().str.title()
            df_xl = df_xl[df_xl['Klient'].str.len() > 1].reset_index(drop=True)

            xl_norm_list = [(normalize(r['Klient']), r) for r in df_xl.to_dict('records')]

            # Match
            df_inv['Buchhalter'] = ''
            df_inv['Pomochnik']  = ''
            for idx, row in df_inv.iterrows():
                b, p = find_accountant(row['Klient'], xl_norm_list)
                df_inv.at[idx, 'Buchhalter'] = b
                df_inv.at[idx, 'Pomochnik']  = p

            # Output path — same folder as CSV
            out_dir  = os.path.dirname(csv_path)
            out_path = os.path.join(out_dir, 'Faktury_po_buchhalterach.xlsx')

            matched, not_found, accountants = generate_excel(df_inv, out_path)

            self.progress.stop()
            self.progress.pack_forget()
            self.status_var.set(f"✓ Gotowe!  Dopasowano: {matched}  |  Nie znaleziono: {not_found}")

            messagebox.showinfo("Gotowe!",
                f"Raport zapisany:\n{out_path}\n\n"
                f"Faktury dopasowane: {matched}\n"
                f"Bez księgowego: {not_found}\n\n"
                f"Księgowi: {', '.join(accountants)}")

        except Exception as e:
            self.progress.stop()
            self.progress.pack_forget()
            self.status_var.set("Błąd!")
            messagebox.showerror("Błąd", f"Wystąpił błąd:\n\n{e}")


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()
